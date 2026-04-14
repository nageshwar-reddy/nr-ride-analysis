#!/usr/bin/env python3
from __future__ import annotations

import datetime
from pathlib import Path
from typing import Dict, Optional, cast
import zipfile

import gpxpy
from gpxpy.gpx import GPX, GPXTrack, GPXTrackSegment
from geopy.distance import distance


# Garmin TrackPointExtension namespace
_TPEX_NS = "{http://www.garmin.com/xmlschemas/TrackPointExtension/v1}"


def _hms(td: datetime.timedelta) -> str:
    """Format a timedelta as "Hh Mm Ss", omitting zero fields."""

    total = int(round(td.total_seconds()))
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    parts: list[str] = []
    if h:
        parts.append(f"{h}h")
    if h or m:
        parts.append(f"{m}m")
    parts.append(f"{s}s")
    return " ".join(parts)


def _to_ist(dt: datetime.datetime) -> str:
    """Convert UTC datetime to IST time string (HH:MM:SS)."""
    ist_offset = datetime.timedelta(hours=5, minutes=30)
    ist_time = dt + ist_offset
    return ist_time.strftime("%H:%M:%S")


def _geocode_location(latitude: float, longitude: float, geocoder, cache: dict) -> str:
    """Reverse geocode with caching. Returns address or 'N/A' on failure."""
    import time

    # Round to ~1km precision for cache key
    cache_key = (round(latitude, 2), round(longitude, 2))

    if cache_key in cache:
        return cache[cache_key]

    try:
        time.sleep(1.1)  # Respect Nominatim rate limit (1 req/sec)
        location = geocoder.reverse(f"{latitude}, {longitude}", timeout=10)
        address = location.address if location else "N/A"
        cache[cache_key] = address
        return address
    except Exception:
        cache[cache_key] = "N/A"
        return "N/A"


def _extract_extension_value(point, tag_suffix: str) -> float | None:
    """Extract a numeric value from gpxtpx TrackPointExtension (hr, cad, pwr, atemp)."""
    for ext in point.extensions:
        tag = getattr(ext, "tag", "")
        if tag.endswith("TrackPointExtension"):
            for child in ext:
                if child.tag == f"{_TPEX_NS}{tag_suffix}":
                    try:
                        return float(child.text)
                    except (ValueError, TypeError):
                        return None
    return None


def _compute_segment_metrics(points: list, min_speed: float) -> dict:
    """Compute riding metrics for a segment (list of GPX trackpoints).

    Returns dict with: seg_avg_speed_kmh, seg_avg_cadence, seg_avg_hr,
    seg_avg_power, seg_elev_gain, seg_elev_loss.
    """
    empty = {
        "seg_avg_speed_kmh": None,
        "seg_avg_cadence": None,
        "seg_avg_hr": None,
        "seg_avg_power": None,
        "seg_elev_gain": 0.0,
        "seg_elev_loss": 0.0,
        "seg_distance_km": 0.0,
        "seg_duration": datetime.timedelta(),
    }
    if len(points) < 2:
        return empty

    total_dist = 0.0
    moving_dist = 0.0
    moving_time = 0.0
    elev_gain = 0.0
    elev_loss = 0.0
    cadence_vals: list[float] = []
    hr_vals: list[float] = []
    power_vals: list[float] = []

    for i in range(len(points)):
        # Extract extensions from every point
        cad = _extract_extension_value(points[i], "cad")
        if cad is not None and cad > 0:
            cadence_vals.append(cad)
        hr = _extract_extension_value(points[i], "hr")
        if hr is not None and hr > 0:
            hr_vals.append(hr)
        pwr = _extract_extension_value(points[i], "pwr")
        if pwr is not None and pwr > 0:
            power_vals.append(pwr)

        if i == 0:
            continue

        prev, curr = points[i - 1], points[i]
        t_prev = getattr(prev, "time", None)
        t_curr = getattr(curr, "time", None)
        if t_prev is None or t_curr is None:
            continue
        dt = (t_curr - t_prev).total_seconds()
        if dt <= 0:
            continue

        d_m = distance((prev.latitude, prev.longitude), (curr.latitude, curr.longitude)).m
        v = d_m / dt
        total_dist += d_m

        if v >= min_speed:
            moving_dist += d_m
            moving_time += dt

        # Elevation deltas
        if prev.elevation is not None and curr.elevation is not None:
            delta = curr.elevation - prev.elevation
            if delta > 0:
                elev_gain += delta
            else:
                elev_loss += abs(delta)

    avg_speed = (moving_dist / moving_time * 3.6) if moving_time > 0 else None

    # Segment duration: wall-clock time from first to last point
    t_first = getattr(points[0], "time", None)
    t_last = getattr(points[-1], "time", None)
    seg_duration = (t_last - t_first) if (t_first and t_last) else datetime.timedelta()

    return {
        "seg_avg_speed_kmh": round(avg_speed, 1) if avg_speed is not None else None,
        "seg_avg_cadence": round(sum(cadence_vals) / len(cadence_vals)) if cadence_vals else None,
        "seg_avg_hr": round(sum(hr_vals) / len(hr_vals)) if hr_vals else None,
        "seg_avg_power": round(sum(power_vals) / len(power_vals)) if power_vals else None,
        "seg_elev_gain": round(elev_gain, 1),
        "seg_elev_loss": round(elev_loss, 1),
        "seg_distance_km": round(total_dist / 1000, 2),
        "seg_duration": seg_duration,
    }


def _fmt_metric(val, unit: str = "") -> str:
    """Format a metric value for display, returning 'N/A' if None."""
    if val is None:
        return "N/A"
    return f"{val}{unit}"


def _print_pause_summary(stats: dict, *, tz_offset: int = 0, enable_geocoding: bool = False) -> None:
    """Human-readable reporting of pause-trimming operation.

    Args:
        stats: dict returned by ``_trim_track`` / ``run_pause_trimmer``.
        tz_offset: Kept for backward compatibility but no longer used.
        enable_geocoding: Whether to perform reverse geocoding for location names.
    """

    if "activity_start" in stats:
        t0 = stats["activity_start"]
    elif stats["pauses"]:
        t0 = min(p["start"] for p in stats["pauses"])
    else:  # no pauses found
        t0 = None

    # Initialize geocoder if enabled
    geocoder = None
    geo_cache = {}
    if enable_geocoding:
        from geopy.geocoders import Nominatim
        geocoder = Nominatim(user_agent="gpx_trimmer")

    print(f"Activity date  {stats['activity_start']:%Y-%m-%d}")
    print(f"Start time  {stats['activity_start']:%H:%M:%S} UTC")
    print(" ")
    print(f"{'Seg':>3}  {'Seg Dist':>10}  {'Seg Dur':>12}  {'Seg End IST':>11}  {'Break Dur':>12}  "
          f"{'Avg Spd':>9}  {'Avg Cad':>7}  {'Avg HR':>6}  "
          f"{'Elev+':>7}  {'Elev-':>7}  {'Avg Pwr':>7}  "
          f"{'Cum.Dist':>10}  "
          f"{'Seg End Rel':>15}  {'Location':>30}  {'Maps Link':>45}  "
          f"{'Drift':>9}")

    for i, p in enumerate(stats["pauses"], 1):
        # Format Δt as HH:MM:SS (zero-padded, hours may exceed 24)
        if t0 is not None:
            rel_td = p["start"] - t0
            total = int(round(rel_td.total_seconds()))
            hh, rem = divmod(total, 3600)
            mm, ss = divmod(rem, 60)
            rel = f"{hh:02d}:{mm:02d}:{ss:02d}"
        else:
            rel = "—"

        gap = _hms(p["gap"])
        cut = _hms(p["removed"])
        drift = f"{round(p['drift']):>3}m"

        # Cumulative distance
        cum_dist = f"{p.get('cumulative_dist', 0)/1000:.2f} km"

        # IST time
        ist_time = _to_ist(p["start"])

        # Segment duration and distance
        seg_dur = _hms(p.get("seg_duration", datetime.timedelta()))
        seg_dist = f"{p.get('seg_distance_km', 0):.2f} km"

        # Location
        if enable_geocoding and geocoder and 'latitude' in p:
            location = _geocode_location(p['latitude'], p['longitude'], geocoder, geo_cache)
            location = location[:28] + ".." if len(location) > 30 else location  # Truncate
        else:
            location = "—"

        # Google Maps link
        if 'latitude' in p and 'longitude' in p:
            maps_link = f"https://maps.google.com/?q={p['latitude']},{p['longitude']}"
        else:
            maps_link = "—"

        # Segment metrics
        spd = _fmt_metric(p.get("seg_avg_speed_kmh"), " km/h")
        cad = _fmt_metric(p.get("seg_avg_cadence"))
        hr = _fmt_metric(p.get("seg_avg_hr"))
        pwr = _fmt_metric(p.get("seg_avg_power"))
        eg = f"{p.get('seg_elev_gain', 0):.0f}m"
        el = f"{p.get('seg_elev_loss', 0):.0f}m"

        print(f"{i:>3}  {seg_dist:>10}  {seg_dur:>12}  {ist_time:>11}  {gap:>12}  "
              f"{spd:>9}  {cad:>7}  {hr:>6}  "
              f"{eg:>7}  {el:>7}  {pwr:>7}  "
              f"{cum_dist:>10}  "
              f"{rel:>15}  {location:>30}  {maps_link:>45}  "
              f"{drift:>9}")

    # Final segment row
    fs = stats.get("final_segment", {})
    if fs:
        seg_num = len(stats["pauses"]) + 1
        cum_dist = f"{stats.get('total_distance', 0)/1000:.2f} km"

        if t0 is not None and "activity_end" in stats:
            rel_td = stats["activity_end"] - t0
            total = int(round(rel_td.total_seconds()))
            hh, rem = divmod(total, 3600)
            mm, ss = divmod(rem, 60)
            rel = f"{hh:02d}:{mm:02d}:{ss:02d}"
        else:
            rel = "—"

        seg_dur = _hms(fs.get("seg_duration", datetime.timedelta()))
        seg_dist = f"{fs.get('seg_distance_km', 0):.2f} km"
        spd = _fmt_metric(fs.get("seg_avg_speed_kmh"), " km/h")
        cad = _fmt_metric(fs.get("seg_avg_cadence"))
        hr = _fmt_metric(fs.get("seg_avg_hr"))
        pwr = _fmt_metric(fs.get("seg_avg_power"))
        eg = f"{fs.get('seg_elev_gain', 0):.0f}m"
        el = f"{fs.get('seg_elev_loss', 0):.0f}m"

        print(f"{seg_num:>3}  {seg_dist:>10}  {seg_dur:>12}  {'Finish':>11}  {'—':>12}  "
              f"{spd:>9}  {cad:>7}  {hr:>6}  "
              f"{eg:>7}  {el:>7}  {pwr:>7}  "
              f"{cum_dist:>10}  "
              f"{rel:>15}  {'—':>30}  {'—':>45}  "
              f"{'—':>9}")

    print(" ")
    print(f"Original elapsed time {_hms(stats['orig_elapsed']):>12}")
    print(f"Trimmed elapsed time {_hms(stats['trimmed_elapsed']):>12}")
    print(f"Total pause time {_hms(stats['removed_time']):>12}")
    print("-" * 55)


def _write_excel_summary(stats: dict, output_path: Path, *, enable_geocoding: bool = False) -> None:
    """Write pause summary to an Excel file.

    Args:
        stats: dict returned by ``_trim_track`` containing pause information.
        output_path: Path where the Excel file should be saved.
        enable_geocoding: Whether geocoding was enabled (for location column).
    """
    import pandas as pd

    if not stats["pauses"]:
        return  # No pauses to write

    # Determine activity start for relative time calculation
    if "activity_start" in stats:
        t0 = stats["activity_start"]
    elif stats["pauses"]:
        t0 = min(p["start"] for p in stats["pauses"])
    else:
        t0 = None

    # Initialize geocoder if needed
    geocoder = None
    geo_cache = {}
    if enable_geocoding:
        from geopy.geocoders import Nominatim
        geocoder = Nominatim(user_agent="gpx_trimmer")

    # Build data rows
    rows = []
    for i, p in enumerate(stats["pauses"], 1):
        # Relative time
        if t0 is not None:
            rel_td = p["start"] - t0
            total = int(round(rel_td.total_seconds()))
            hh, rem = divmod(total, 3600)
            mm, ss = divmod(rem, 60)
            rel_time = f"{hh:02d}:{mm:02d}:{ss:02d}"
        else:
            rel_time = "—"

        # Duration and removed time
        gap_seconds = p["gap"].total_seconds()

        # Formatted duration for display
        duration_formatted = _hms(p["gap"])

        # Cumulative distance
        cum_dist_km = p.get('cumulative_dist', 0) / 1000

        # IST time
        ist_time = _to_ist(p["start"])

        # Location
        location = "—"
        if enable_geocoding and geocoder and 'latitude' in p:
            location = _geocode_location(p['latitude'], p['longitude'], geocoder, geo_cache)

        # Google Maps link
        maps_link = ""
        if 'latitude' in p and 'longitude' in p:
            maps_link = f"https://maps.google.com/?q={p['latitude']},{p['longitude']}"

        rows.append({
            "Segment #": i,
            "Segment Distance (km)": p.get("seg_distance_km", 0),
            "Segment Duration": _hms(p.get("seg_duration", datetime.timedelta())),
            "Seg. End IST": ist_time,
            "Break Duration": duration_formatted,
            "Avg Speed (km/h)": p.get("seg_avg_speed_kmh") if p.get("seg_avg_speed_kmh") is not None else "N/A",
            "Avg Cadence": p.get("seg_avg_cadence") if p.get("seg_avg_cadence") is not None else "N/A",
            "Avg HR": p.get("seg_avg_hr") if p.get("seg_avg_hr") is not None else "N/A",
            "Elev Gain (m)": p.get("seg_elev_gain", 0),
            "Elev Loss (m)": p.get("seg_elev_loss", 0),
            "Avg Power (W)": p.get("seg_avg_power") if p.get("seg_avg_power") is not None else "N/A",
            "Cumulative Distance (km)": round(cum_dist_km, 2),
            "Seg. End Relative": rel_time,
            "Location": location,
            "Google Maps Link": maps_link,
            "Break Duration (seconds)": gap_seconds,
            "Drift (meters)": round(p['drift']),
            "Latitude": p.get('latitude', ''),
            "Longitude": p.get('longitude', ''),
        })

    # Final segment row
    fs = stats.get("final_segment", {})
    if fs:
        seg_num = len(stats["pauses"]) + 1
        cum_dist_km = stats.get("total_distance", 0) / 1000

        if t0 is not None and "activity_end" in stats:
            rel_td = stats["activity_end"] - t0
            total = int(round(rel_td.total_seconds()))
            hh, rem = divmod(total, 3600)
            mm, ss = divmod(rem, 60)
            rel_time = f"{hh:02d}:{mm:02d}:{ss:02d}"
        else:
            rel_time = "—"

        rows.append({
            "Segment #": seg_num,
            "Segment Distance (km)": fs.get("seg_distance_km", 0),
            "Segment Duration": _hms(fs.get("seg_duration", datetime.timedelta())),
            "Seg. End IST": "Finish",
            "Break Duration": "—",
            "Avg Speed (km/h)": fs.get("seg_avg_speed_kmh") if fs.get("seg_avg_speed_kmh") is not None else "N/A",
            "Avg Cadence": fs.get("seg_avg_cadence") if fs.get("seg_avg_cadence") is not None else "N/A",
            "Avg HR": fs.get("seg_avg_hr") if fs.get("seg_avg_hr") is not None else "N/A",
            "Elev Gain (m)": fs.get("seg_elev_gain", 0),
            "Elev Loss (m)": fs.get("seg_elev_loss", 0),
            "Avg Power (W)": fs.get("seg_avg_power") if fs.get("seg_avg_power") is not None else "N/A",
            "Cumulative Distance (km)": round(cum_dist_km, 2),
            "Seg. End Relative": rel_time,
            "Location": "—",
            "Google Maps Link": "",
            "Break Duration (seconds)": "—",
            "Drift (meters)": "—",
            "Latitude": stats.get("end_latitude", ""),
            "Longitude": stats.get("end_longitude", ""),
        })

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Build column name → 1-based index mapping for styling
    col_index = {name: idx + 1 for idx, name in enumerate(df.columns)}

    # Add summary information as separate sheet data
    summary_data = {
        "Metric": [
            "Activity Date",
            "Start Time (UTC)",
            "Original Elapsed Time",
            "Trimmed Elapsed Time",
            "Total Pause Time Removed"
        ],
        "Value": [
            stats['activity_start'].strftime('%Y-%m-%d'),
            stats['activity_start'].strftime('%H:%M:%S'),
            _hms(stats['orig_elapsed']),
            _hms(stats['trimmed_elapsed']),
            _hms(stats['removed_time'])
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # Write to Excel with multiple sheets
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Pauses', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Define styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        alt_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Column indices by name
        ci_dur_sec = col_index["Break Duration (seconds)"]
        ci_maps = col_index["Google Maps Link"]
        center_cols = {
            col_index["Segment #"],
            col_index["Seg. End IST"], col_index["Segment Duration"],
            col_index["Segment Distance (km)"], col_index["Break Duration"],
            col_index["Drift (meters)"], col_index["Cumulative Distance (km)"],
            col_index["Break Duration (seconds)"], col_index["Seg. End Relative"],
            col_index["Avg Speed (km/h)"], col_index["Avg Cadence"],
            col_index["Avg HR"], col_index["Avg Power (W)"],
            col_index["Elev Gain (m)"], col_index["Elev Loss (m)"],
        }

        # Format Pauses sheet
        ws_pauses = writer.sheets['Pauses']

        # Style header row
        for col_idx, cell in enumerate(ws_pauses[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

            # Add note to Duration (seconds) header
            if col_idx == ci_dur_sec:
                cell.font = Font(bold=True, color='FFFFFF', size=9, italic=True)
                from openpyxl.comments import Comment
                cell.comment = Comment("Numeric value for pivot tables and analysis", "GPX Trimmer")

        # Style data rows
        for row_idx, row in enumerate(ws_pauses.iter_rows(min_row=2, max_row=len(df)+1), start=2):
            # Alternating row colors
            if row_idx % 2 == 0:
                fill = alt_fill
            else:
                fill = PatternFill(fill_type=None)

            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.fill = fill

                if col_idx in center_cols:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

                # Make Duration (seconds) column subtle
                if col_idx == ci_dur_sec:
                    cell.font = Font(color='808080', size=9)

                # Format Google Maps Link as hyperlink
                if col_idx == ci_maps and cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color='0563C1', underline='single')
                    cell.value = 'View on Maps'

        # Auto-adjust column widths
        for col_idx, column in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(column)) + 2
            for cell in ws_pauses[column_letter][1:]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_pauses.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Freeze header row
        ws_pauses.freeze_panes = 'A2'

        # Format Summary sheet
        ws_summary = writer.sheets['Summary']

        # Style header row
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Style data rows
        for row_idx, row in enumerate(ws_summary.iter_rows(min_row=2, max_row=len(summary_df)+1), start=2):
            if row_idx % 2 == 0:
                fill = alt_fill
            else:
                fill = PatternFill(fill_type=None)

            for cell in row:
                cell.border = border
                cell.fill = fill
                cell.alignment = left_align
                cell.font = Font(size=10)

        # Auto-adjust column widths for summary
        for col_idx in range(1, len(summary_df.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws_summary[column_letter]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_summary.column_dimensions[column_letter].width = max_length + 4

        # Make Metric column bold in summary
        for row in ws_summary.iter_rows(min_row=2, max_row=len(summary_df)+1, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True, size=10)

    print(f"Excel summary saved: {output_path.name}")


def _decode_name(info: zipfile.ZipInfo) -> str:
    """Return a proper string for *info.filename*.

    Some zips lack the UTF-8 flag; repair their names.
    """
    # bit 11 set → filename is already UTF-8
    if info.flag_bits & 0x800:
        return info.filename
    raw = info.filename.encode("cp437")  # undo zipfile's default decoding
    try:
        return raw.decode("utf-8")  # most likely correct
    except UnicodeDecodeError:
        return raw.decode("latin-1")  # graceful fallback


def _ts(t: Optional[datetime.datetime]) -> datetime.datetime:
    """Return *t* if it's a datetime, else raise."""
    if t is None:
        raise ValueError("GPX point is missing its <time> stamp")
    return cast(datetime.datetime, t)


def _trim_track(original: GPX, *, min_speed: float = 0.5, min_pause_duration: int = 600) -> Dict:
    """
    Detect long, low-speed pauses in a GPX track and compute per-segment
    riding metrics (speed, cadence, heart rate, power, elevation).

    A *segment* is the riding portion between consecutive pauses.

    Args:
        original: The parsed GPX object to analyse.
        min_speed: Speed threshold, in m s-1, below which motion is considered
            stationary.
        min_pause_duration: Minimum pause length, in seconds, that must be sustained
            before it is recorded.

    Returns:
        stats: A dictionary containing
            * ``pauses``          - list of pause dicts (each with segment metrics)
            * ``removed_time``    - total pause time as ``timedelta``
            * ``pause_drift``     - total "drift" distance in metres
            * ``orig_elapsed``    - elapsed time before trimming
            * ``trimmed_elapsed`` - elapsed time after trimming
            * ``final_segment``   - metrics dict for the last riding segment
            * ``total_distance``  - total cumulative distance in metres
            * ``activity_start``  - start datetime
            * ``activity_end``    - end datetime

    Notes:
        soft pause: *Sequence of points within ONE segment* whose instantaneous speed
            drops below `min_speed`. We keep just enough of the pause to cover the drift
            distance at the *average moving speed so far* and cut the rest.
        hard pause: Gap between two consecutive segments. Same rule: we keep the minimum time
            needed to traverse the straight-line distance at average speed and trim the excess.
        Returned `stats["pauses"]` rows therefore show gap >= removed >= 0  for **both** pause kinds.
    """
    stats = {  # aggregate totals + pause list
        "pauses": [],  # list[dict]
        "removed_time": datetime.timedelta(),
        "pause_drift": 0.0,
    }

    cumulative_dist = 0.0  # total distance traveled
    segment_points: list = []  # trackpoints in current riding segment

    # ────────────────────────── iterate over tracks & segments ──────────────
    for src_trk in original.tracks:
        moving_dist = moving_time = 0.0  # for average-speed estimates

        for seg_idx, src_seg in enumerate(src_trk.segments):
            if not src_seg.points:
                continue

            segment_points.append(src_seg.points[0])

            # state for a potential soft pause
            p_start_idx = None  # first low-speed pt index
            p_start_time = None  # timestamp of preceding pt
            p_drift = 0.0  # metres drifted during pause

            pts = src_seg.points
            for i in range(1, len(pts)):
                prev, curr = pts[i - 1], pts[i]
                dt = (_ts(curr.time) - _ts(prev.time)).total_seconds()
                if dt <= 0:  # duplicate / rewind in source
                    segment_points.append(curr)
                    continue

                # instantaneous speed between two source points
                d_m = distance((prev.latitude, prev.longitude), (curr.latitude, curr.longitude)).m
                v = d_m / dt

                # ── LOW-SPEED block ──────────────────────────────────────
                if v < min_speed:  # inside a soft pause
                    if p_start_idx is None:  # first time we dip below v_min
                        p_start_idx, p_start_time = i, prev.time
                        p_drift = 0.0
                    p_drift += d_m
                    continue

                # ── LEAVING a soft pause ────────────────────────────────
                if p_start_idx is not None:
                    gap = _ts(curr.time) - _ts(p_start_time)

                    if gap.total_seconds() >= min_pause_duration:
                        # amount of that gap we must PRESERVE to maintain speed
                        v_avg = moving_dist / moving_time if moving_time else 0.0
                        keep = p_drift / v_avg if v_avg else 1.0
                        keep = min(keep, gap.total_seconds())  # never > gap
                        cut = gap - datetime.timedelta(seconds=keep)

                        # compute segment metrics before recording pause
                        seg_metrics = _compute_segment_metrics(segment_points, min_speed)

                        # record stats
                        stats["pauses"].append(dict(
                            start=p_start_time,
                            gap=gap,
                            removed=cut,
                            drift=p_drift,
                            cumulative_dist=cumulative_dist,
                            latitude=prev.latitude,
                            longitude=prev.longitude,
                            **seg_metrics,
                        ))
                        stats["removed_time"] += cut
                        stats["pause_drift"] += p_drift

                        # start new riding segment
                        segment_points = []
                    else:  # pause too short → keep intact, add points back
                        for j in range(p_start_idx, i):
                            segment_points.append(pts[j])

                    p_start_idx = p_start_time = None
                    p_drift = 0.0

                # point is normal moving data
                segment_points.append(curr)
                moving_dist += d_m
                moving_time += dt
                cumulative_dist += d_m

            # ── SOFT pause that reaches end of segment ──────────────────
            if p_start_idx is not None:
                gap = _ts(pts[-1].time) - _ts(p_start_time)
                if gap.total_seconds() >= min_pause_duration:
                    v_avg = moving_dist / moving_time if moving_time else 0.0
                    keep = p_drift / v_avg if v_avg else 1.0
                    keep = min(keep, gap.total_seconds())
                    cut = gap - datetime.timedelta(seconds=keep)

                    seg_metrics = _compute_segment_metrics(segment_points, min_speed)

                    stats["pauses"].append(dict(
                        start=p_start_time,
                        gap=gap,
                        removed=cut,
                        drift=p_drift,
                        cumulative_dist=cumulative_dist,
                        latitude=pts[p_start_idx - 1].latitude,
                        longitude=pts[p_start_idx - 1].longitude,
                        **seg_metrics,
                    ))
                    stats["removed_time"] += cut
                    stats["pause_drift"] += p_drift

                    segment_points = []
                else:  # short pause → keep
                    for j in range(p_start_idx, len(pts)):
                        segment_points.append(pts[j])

            # ── HARD pause (gap between segments) ───────────────────────
            nxt = seg_idx + 1
            if nxt < len(src_trk.segments) and src_trk.segments[nxt].points:
                last_pt = src_seg.points[-1]
                first_nx = src_trk.segments[nxt].points[0]
                dt_gap = (_ts(first_nx.time) - _ts(last_pt.time)).total_seconds()
                if dt_gap >= min_pause_duration:
                    d_gap = distance((last_pt.latitude, last_pt.longitude), (first_nx.latitude, first_nx.longitude)).m
                    v_avg = moving_dist / moving_time if moving_time else 0.0
                    keep = d_gap / v_avg if v_avg else 1.0
                    keep = min(keep, dt_gap)
                    cut = datetime.timedelta(seconds=dt_gap - keep)

                    seg_metrics = _compute_segment_metrics(segment_points, min_speed)

                    stats["pauses"].append(dict(
                        start=last_pt.time,
                        gap=datetime.timedelta(seconds=dt_gap),
                        removed=cut,
                        drift=d_gap,
                        cumulative_dist=cumulative_dist,
                        latitude=last_pt.latitude,
                        longitude=last_pt.longitude,
                        **seg_metrics,
                    ))
                    stats["removed_time"] += cut
                    stats["pause_drift"] += d_gap

                    segment_points = []

    # ── overall elapsed times ───────────────────────────────────────────
    stats["activity_start"] = _ts(original.tracks[0].segments[0].points[0].time)
    stats["activity_end"] = _ts(original.tracks[-1].segments[-1].points[-1].time)
    stats["orig_elapsed"] = stats["activity_end"] - stats["activity_start"]
    stats["trimmed_elapsed"] = stats["orig_elapsed"] - stats["removed_time"]
    stats["total_distance"] = cumulative_dist

    # End point coordinates (for final segment row)
    last_pt = original.tracks[-1].segments[-1].points[-1]
    stats["end_latitude"] = last_pt.latitude
    stats["end_longitude"] = last_pt.longitude

    # Final segment metrics (riding after the last pause)
    stats["final_segment"] = _compute_segment_metrics(segment_points, min_speed)

    return stats


def run_pause_trimmer(
    input_path: str | Path,
    *,
    min_speed: float = 0.1,
    min_pause_duration: int = 240,
    enable_geocoding: bool = False,
) -> None:
    """
    Analyse every GPX track in *input_path* for pauses and segment metrics.

    Args:
        input_path: Either a single ``.gpx`` file or a ``.zip`` containing many GPX
            files (any sub-folder layout is preserved).
        min_speed : Low-speed threshold in m s-1 (default 0.1).
        min_pause_duration : Minimum pause duration in seconds before it is recorded (default 240).
        enable_geocoding : Enable reverse geocoding to show location names (default False).
    """
    input_path = Path(input_path)

    # ── helper for one GPX blob ────────────────────────────────────
    def _analyse_and_report(xml: str, label: str) -> dict:
        gpx = gpxpy.parse(xml)
        stats = _trim_track(gpx, min_speed=min_speed, min_pause_duration=min_pause_duration)

        print(f"\n=== {label} ===\n")
        _print_pause_summary(stats, tz_offset=0, enable_geocoding=enable_geocoding)
        return stats

    # ── single GPX on disk ────────────────────────────────────────
    if input_path.suffix.lower() != ".zip":
        xml_in = input_path.read_text(encoding="utf-8", errors="replace")
        stats = _analyse_and_report(xml_in, input_path.name)

        # Write Excel summary
        excel_file = input_path.with_suffix(".xlsx")
        _write_excel_summary(stats, excel_file, enable_geocoding=enable_geocoding)

        return

    # ── ZIP archive ───────────────────────────────────────────────
    with zipfile.ZipFile(input_path) as zin:
        gpx_count = 0

        # walk all entries
        for member in zin.infolist():
            arcname = _decode_name(member)  # repaired text
            p = Path(arcname)

            # skip non-GPX or macOS "resource-fork" files
            if p.suffix.lower() != ".gpx" or p.name.startswith("._"):
                continue

            xml_in = zin.read(member).decode("utf-8", errors="replace")
            stats = _analyse_and_report(xml_in, p.name)

            # Write Excel summary for this GPX file
            excel_file = input_path.parent / f"{p.stem}.xlsx"
            _write_excel_summary(stats, excel_file, enable_geocoding=enable_geocoding)
            gpx_count += 1

    if gpx_count:
        print(f"\nProcessed {gpx_count} GPX file(s).")
    else:
        print("No .gpx files found in the archive.")


if __name__ == "__main__":
    """Main entry point for command-line usage."""

    import argparse

    parser = argparse.ArgumentParser(prog="gpx_trimmer")
    parser.add_argument(
        "--min_speed",
        default=0.1,
        type=float,
        help="Minimum speed in m/s; points below this are considered part of a pause.",
    )
    parser.add_argument(
        "--min_pause_duration",
        default=240,
        type=int,
        help="Minimum pause duration in seconds; pauses longer than this will be detected.",
    )
    parser.add_argument(
        "--enable_geocoding",
        action="store_true",
        help="Enable reverse geocoding to show location names (requires internet).",
    )
    parser.add_argument("input_file_path", help="Input file path")
    args = parser.parse_args()

    run_pause_trimmer(
        args.input_file_path,
        min_speed=args.min_speed,
        min_pause_duration=args.min_pause_duration,
        enable_geocoding=args.enable_geocoding
    )
